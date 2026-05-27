import pandas as pd
import numpy as np
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from analysis import load_data, compute_summary, compute_phase_averages

DARK      = "0B0B0D"
ACCENT    = "2563EB"
GREEN     = "16A34A"
RED       = "DC2626"
YELLOW    = "D97706"
ORANGE    = "EA580C"
WHITE     = "F8FAFC"
MID       = "94A3B8"
HEADER_BG = "1E293B"

def border():
    s = Side(border_style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(color):  return PatternFill("solid", fgColor=color)
def bw(sz=11):    return Font(bold=True, color=WHITE, size=sz)
def center():     return Alignment(horizontal="center", vertical="center")
def cw(ws,col,w): ws.column_dimensions[get_column_letter(col)].width = w

def header_row(ws, row, cols, texts):
    for col, txt in zip(cols, texts):
        c = ws.cell(row=row, column=col, value=txt)
        c.font = bw(); c.fill = fill(HEADER_BG)
        c.alignment = center(); c.border = border()

def sheet_summary(wb, summary, df):
    ws = wb.create_sheet("Inflation Summary")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"].value = "India CPI Inflation Analysis — Jan 2019 to Dec 2025"
    ws["A1"].font = Font(bold=True, color=WHITE, size=14)
    ws["A1"].fill = fill(DARK); ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    ws["A2"].value = "Source: MoSPI CPI All-India (Base Year 2012 = 100)  |  RBI Upper Tolerance Band: 6%"
    ws["A2"].font = Font(color=MID, size=10, italic=True)
    ws["A2"].fill = fill("0F172A"); ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 18

    # Key stats block
    stats = [
        ("Peak General Inflation", f"{summary['peak_general_inflation_pct']}%", summary['peak_general_inflation_date'], RED),
        ("Peak Food Inflation",    f"{summary['peak_food_inflation_pct']}%",    summary['peak_food_inflation_date'],    ORANGE),
        ("Peak Fuel Inflation",    f"{summary['peak_fuel_inflation_pct']}%",    summary['peak_fuel_inflation_date'],    YELLOW),
        ("Total CPI Rise (2019–2025)", f"+{summary['total_cpi_rise_pct']}%",   "Jan 2019 → Dec 2025",                  RED),
        ("Purchasing Power of ₹100",   f"₹{summary['purchasing_power_dec2025']}", "by Dec 2025",                       ORANGE),
        ("Months Above RBI 6% Band",   str(summary['months_above_6pct']),       "out of 72 months",                    YELLOW),
    ]

    for i, (label, value, note, color) in enumerate(stats):
        row = 4 + i
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(bold=True, size=11, color="1E293B")
        lc.fill = fill("F1F5F9"); lc.alignment = center(); lc.border = border()
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = Font(bold=True, size=13, color=color)
        vc.fill = fill(WHITE); vc.alignment = center(); vc.border = border()
        nc = ws.cell(row=row, column=3, value=note)
        nc.font = Font(size=10, color="64748B", italic=True)
        nc.fill = fill(WHITE); nc.alignment = center(); nc.border = border()
        ws.row_dimensions[row].height = 22

    # Monthly YoY table
    header_row(ws, 11, range(1,7),
               ["Date","General %","Food %","Fuel %","Housing %","Purchasing Power ₹"])
    ws.row_dimensions[11].height = 20

    df_full = df.dropna(subset=["cpi_general_yoy_pct"]).reset_index(drop=True)
    for ri, row in enumerate(df_full.itertuples(), start=12):
        bg = fill("F1F5F9") if ri%2==0 else fill(WHITE)
        vals = [
            row.date.strftime("%b %Y"),
            row.cpi_general_yoy_pct,
            row.cpi_food_yoy_pct,
            row.cpi_fuel_yoy_pct,
            row.cpi_housing_yoy_pct,
            row.purchasing_power_100,
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.fill = bg; c.alignment = center(); c.border = border()
            c.font = Font(size=10, color="1E293B")
            if col == 2 and isinstance(val, float):
                c.font = Font(size=10, bold=True,
                              color=RED if val>6 else (YELLOW if val>4 else GREEN))
        ws.row_dimensions[ri].height = 16

    for col, w in zip(range(1,7), [12,13,10,10,12,18]):
        cw(ws, col, w)

def sheet_phases(wb, phases):
    ws = wb.create_sheet("Phase Analysis")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"].value = "Inflation by Market Phase — Average YoY % per Category"
    ws["A1"].font = Font(bold=True, color=WHITE, size=13)
    ws["A1"].fill = fill(DARK); ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 30

    header_row(ws, 2, range(1,6),
               ["Phase","Avg General %","Avg Food %","Avg Fuel %","Months > 6%"])
    ws.row_dimensions[2].height = 20

    for ri, row in enumerate(phases.itertuples(), start=3):
        bg = fill("F1F5F9") if ri%2==0 else fill(WHITE)
        vals = [row.phase, row.avg_general_inflation,
                row.avg_food_inflation, row.avg_fuel_inflation, row.months_above_6pct]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.fill = bg; c.alignment = center(); c.border = border()
            c.font = Font(size=10, color="1E293B")
            if col == 2 and isinstance(val, float):
                c.font = Font(size=10, bold=True,
                              color=RED if val>6 else (YELLOW if val>4 else GREEN))
        ws.row_dimensions[ri].height = 20

    last = 2 + len(phases)
    chart = BarChart()
    chart.type = "col"
    chart.title = "Average Inflation by Phase"
    chart.y_axis.title = "YoY %"
    chart.style = 10; chart.width = 22; chart.height = 12
    chart.add_data(Reference(ws, min_col=2, max_col=4, min_row=2, max_row=last), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=3, max_row=last))
    ws.add_chart(chart, f"A{last+3}")

    for col, w in zip(range(1,6), [24,16,12,12,14]):
        cw(ws, col, w)

def sheet_purchasing_power(wb, df):
    ws = wb.create_sheet("Purchasing Power")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    ws["A1"].value = "Erosion of Purchasing Power — ₹100 in Jan 2019"
    ws["A1"].font = Font(bold=True, color=WHITE, size=13)
    ws["A1"].fill = fill(DARK); ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 30

    header_row(ws, 2, range(1,5), ["Date","CPI General","Real Value of ₹100","Erosion %"])
    ws.row_dimensions[2].height = 20

    for ri, row in enumerate(df.itertuples(), start=3):
        bg = fill("F1F5F9") if ri%2==0 else fill(WHITE)
        vals = [row.date.strftime("%b %Y"), row.cpi_general,
                row.purchasing_power_100, row.real_erosion_pct]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.fill = bg; c.alignment = center(); c.border = border()
            c.font = Font(size=10, color="1E293B")
            if col == 4 and isinstance(val, float):
                c.font = Font(size=10, bold=True,
                              color=RED if val>30 else (ORANGE if val>20 else YELLOW))
        ws.row_dimensions[ri].height = 16

    for col, w in zip(range(1,5), [12,14,20,14]):
        cw(ws, col, w)

if __name__ == "__main__":
    df = load_data()
    summary = compute_summary(df)
    phases = compute_phase_averages(df)

    wb = Workbook()
    wb.remove(wb.active)
    sheet_summary(wb, summary, df)
    sheet_phases(wb, phases)
    sheet_purchasing_power(wb, df)
    wb.save("inflation_cost_of_living_dashboard.xlsx")
    print("Saved → inflation_cost_of_living_dashboard.xlsx")